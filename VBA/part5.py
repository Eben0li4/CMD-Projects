import openpyxl
import time

# Open the source workbook and select the worksheet
source_wb = openpyxl.load_workbook(r"C:\Users\EbenOlivier\OneDrive - 4 Arrows Mining\Reporting\Data2(Speed up OneDrive)\KMRC Availability Per Day.xlsx")
source_ws = source_wb["KMRC Availability Per Day"]

# Open the destination workbook and select the worksheet
dest_wb = openpyxl.load_workbook("C:\Dropbox\Intellicode Production Reports\Report Templates\KMR Operations Report.xlsx")
dest_ws = dest_wb["Availability"]

# Copy data from source worksheet to destination worksheet
for row in source_ws.iter_rows(min_row=1, max_row=1499, min_col=1, max_col=21):
    for cell in row:
        dest_ws.cell(row=cell.row + 6012, column=cell.column).value = cell.value

# Save the destination workbook
dest_wb.save("C:\Dropbox\Intellicode Production Reports\Report Templates\KMR Operations Report.xlsx")

# Wait for 3 seconds
time.sleep(3)
