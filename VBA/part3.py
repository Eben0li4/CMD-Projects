import openpyxl

# Open the source workbook and get the source worksheet
source_path = "C:\\Users\\EbenOlivier\\OneDrive - 4 Arrows Mining\\Reporting\\Data\\Crushing Dump KMRC.xlsx"
source_wb = openpyxl.load_workbook(source_path, data_only=True)
source_ws = source_wb["Crushing Dump KMRC"]

# Read the data from the source worksheet into a list
data = [[cell.value for cell in row] for row in source_ws.iter_rows(min_row=1, max_row=2000, min_col=1, max_col=33)]

# Open the destination workbook and get the destination worksheet
destination_path = "C:\\Dropbox\\Intellicode Production Reports\\Report Templates\\KMR Operations Report.xlsx"
destination_wb = openpyxl.load_workbook(destination_path)
destination_ws = destination_wb["CrushingData"]

# Write the data to the destination worksheet
for row_index, row in enumerate(data, start=1):
    for col_index, value in enumerate(row, start=1):
        destination_ws.cell(row=row_index, column=col_index, value=value)

# Save the destination workbook
destination_wb.save(destination_path)

# Close the workbooks
source_wb.close()
destination_wb.close()
