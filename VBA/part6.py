from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import pandas as pd
import re

# Function to copy data from one worksheet to another
def copy_data(source_path, source_sheet, source_range, dest_workbook, dest_sheet, dest_start_cell):
    source_wb = load_workbook(source_path, data_only=True)
    source_ws = source_wb[source_sheet]
    
    # Extract data from the specified range
    data = []
    for row in source_ws[source_range]:
        data.append([cell.value for cell in row])
    
    # Convert the data to a DataFrame
    data_df = pd.DataFrame(data)
    
    dest_ws = dest_workbook[dest_sheet]
    
    # Parse the destination start cell
    match = re.match(r"([A-Z]+)(\d+)", dest_start_cell)
    start_col, start_row = match.groups()
    start_col = column_index_from_string(start_col)
    start_row = int(start_row)
    
    # Start writing data at the specified start cell
    for r_idx, row in enumerate(data_df.values, start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            dest_ws.cell(row=r_idx, column=c_idx, value=value)
    
    source_wb.close()

# Open the destination workbook
dest_path = r"C:\Dropbox\Intellicode Production Reports\Report Templates\KMR Crushing Report.xlsx"
dest_wb = load_workbook(dest_path)

# Copy data from KMRC Availability Per Day.xlsx
copy_data(r"C:\Users\EbenOlivier\OneDrive - 4 Arrows Mining\Reporting\Data2(Speed up OneDrive)\KMRC Availability Per Day.xlsx",
          "KMRC Availability Per Day", "A1:U1499", dest_wb, "Availability", "A1")

# Copy data from 4AM Current Plantlist.xlsx
copy_data("C:/Users/EbenOlivier/OneDrive - 4 Arrows Mining/Reporting/Data2(Speed up OneDrive)/4AM Current Plantlist.xlsx",
          "4AM Current Plantlist", "A1:N2000", dest_wb, "Fleetlist", "A1")

# Save the destination workbook
dest_wb.save(dest_path)

# Export the 'PlantReport' sheet to PDF
import win32com.client as win32

excel = win32.Dispatch('Excel.Application')
wb = excel.Workbooks.Open(dest_path)
ws = wb.Worksheets("PlantReport")
pdf_path = "C:/Dropbox/Intellicode Production Reports/Reports/KMR Crushing Plant Report.pdf"
ws.ExportAsFixedFormat(0, pdf_path)

wb.Close(True)
excel.Quit()
