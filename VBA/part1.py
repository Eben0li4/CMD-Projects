import openpyxl

# Disable screen updating and alerts
# This is not directly applicable in Python as it typically runs in non-interactive mode
# and doesn't have a concept of screen updating or alerts like VBA does.

# Transform Data Files

# Find & replace in KMRC Availability Per Day.xlsx
kmrc_availability_path = r"C:\Users\EbenOlivier\OneDrive - 4 Arrows Mining\Reporting\Data\KMRC Availability Per Day.xlsx"
kmrc_availability_wb = openpyxl.load_workbook(kmrc_availability_path)
ws_availability = kmrc_availability_wb["KMRC Availability Per Day"]

for cell in ws_availability['L']:
    if cell.value == "-33":
        cell.value = 0

kmrc_availability_wb.save(kmrc_availability_path)
kmrc_availability_wb.close()

# Change text to number \ date format of BD Dash Data
kmrc_breakdown_path = r"C:\Users\EbenOlivier\OneDrive - 4 Arrows Mining\Reporting\Data\KMRC Breakdown Dash.xlsx"
kmrc_breakdown_wb = openpyxl.load_workbook(kmrc_breakdown_path)
ws_breakdown = kmrc_breakdown_wb.active

ws_breakdown['M1'] = 1

for row in ws_breakdown.iter_rows(min_row=2, max_row=1000, min_col=6, max_col=12):
    for cell in row:
        if isinstance(cell.value, str) and cell.value.isdigit():
            cell.value = int(cell.value)

kmrc_breakdown_wb.save(kmrc_breakdown_path)
kmrc_breakdown_wb.close()
