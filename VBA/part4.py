import openpyxl

# Open the Excel files
crushing_report_path = "C:\\Dropbox\\Intellicode Production Reports\\Report Templates\\KMR Crushing Report.xlsx"
operations_report_path = "C:\\Dropbox\\Intellicode Production Reports\\Report Templates\\KMR Operations Report.xlsx"

crushing_wb = openpyxl.load_workbook(crushing_report_path)
operations_wb = openpyxl.load_workbook(operations_report_path)

# Access the specific sheets
crushing_sheet = crushing_wb["CrushingData"]
operations_sheet = operations_wb["CrushingData"]

# Copy the data from the Crushing Report to the Operations Report
for row in range(2004, 3001):
    for col in range(1, 34):  # A:AG is 1:33 in 1-indexed Excel columns
        operations_sheet.cell(row=row, column=col).value = crushing_sheet.cell(row=row, column=col).value

for row in range(2013, 3001):
    operations_sheet.cell(row=row, column=36).value = crushing_sheet.cell(row=row, column=38).value  # AJ is 36 and AL is 38

# Save the changes to the Operations Report
operations_wb.save(operations_report_path)

# Close the Crushing Report without saving changes
crushing_wb.close()
